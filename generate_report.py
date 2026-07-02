#!/usr/bin/env python3
"""
Generate an Excel report of a store's Printful sync products and their artwork
status (real uploaded artwork vs. parametric text/clipart designs).

Produces two sheets:
  * Products  — one row per sync product with technique(s), layer types, a
                color-coded "Real Artwork?" flag, and thumbnail URL.
  * Variants  — every sync variant with SKU, catalog variant id, retail price,
                technique, layer types, and any artwork URL.

The workbook is intentionally formula-free (summary cells are computed here) so
it displays correct values without a LibreOffice/Excel recalculation pass.

Requires openpyxl:  pip install "printful-mcp[reports]"  (or: pip install openpyxl)

Usage:
    PRINTFUL_API_KEY=... PRINTFUL_STORE_ID=... python generate_report.py [store_id] [out.xlsx]
"""
import json
import os
import sys
import urllib.request
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_KEY = os.environ["PRINTFUL_API_KEY"]
STORE = sys.argv[1] if len(sys.argv) > 1 else os.environ["PRINTFUL_STORE_ID"]
OUT = sys.argv[2] if len(sys.argv) > 2 else "sync_products_report.xlsx"

FONT = "Arial"
NAVY = "1F3864"


def get(url):
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + API_KEY, "X-PF-Store-Id": STORE})
    return json.load(urllib.request.urlopen(req))


def fetch():
    products = get("https://api.printful.com/v2/sync-products?limit=100")["data"]
    prod_rows, var_rows = [], []
    for p in products:
        variants = get(f"https://api.printful.com/v2/sync-products/{p['id']}/sync-variants?limit=100")["data"]
        techs, ltypes, urls = set(), set(), set()
        for v in variants:
            placements = v.get("placements", [])
            for pl in placements:
                techs.add(pl.get("technique"))
                for l in pl.get("layers", []):
                    ltypes.add(l.get("type"))
                    if l.get("url"):
                        urls.add(l["url"])
            var_rows.append([
                p["name"], v.get("name"), v.get("sku"), v.get("catalog_variant_id"),
                float(v["retail_price"]) if v.get("retail_price") else None, v.get("currency"),
                ", ".join(sorted({pl.get("technique") for pl in placements if pl.get("technique")})),
                ", ".join(sorted({l.get("type") for pl in placements for l in pl.get("layers", []) if l.get("type")})),
                next(iter(l["url"] for pl in placements for l in pl.get("layers", []) if l.get("url")), ""),
            ])
        prod_rows.append([
            p["name"], p["id"], p.get("external_id"), len(variants),
            ", ".join(sorted(t for t in techs if t)), ", ".join(sorted(t for t in ltypes if t)),
            "Yes" if urls else "No", p.get("thumbnail_url", ""),
        ])
    return prod_rows, var_rows


def write_sheet(ws, title, headers, rows, widths, artwork_col=None):
    hdr_fill = PatternFill("solid", start_color=NAVY)
    hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    base_font = Font(name=FONT, size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green = Font(name=FONT, size=10, color="008000", bold=True)
    red = Font(name=FONT, size=10, color="C00000", bold=True)

    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, bold=True, size=14, color=NAVY)
    hrow = 3
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hrow, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for r, row in enumerate(rows, hrow + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.font = base_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if artwork_col and c == artwork_col:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = green if val == "Yes" else red
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(hrow + 1, 1)
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(headers))}{hrow + len(rows)}"
    return hrow + len(rows)


def build(prod_rows, var_rows):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Products"
    ph = ["Product", "Sync ID", "External ID", "Variants", "Technique(s)", "Layer Types", "Real Artwork?", "Thumbnail URL"]
    last = write_sheet(ws1, "Printful Sync Products — Artwork Scan", ph, prod_rows,
                       [40, 12, 15, 10, 16, 18, 14, 60], artwork_col=7)

    n_art = sum(1 for r in prod_rows if r[6] == "Yes")
    srow = last + 2
    ws1.cell(srow, 1, "Products with real artwork:").font = Font(name=FONT, bold=True, size=10)
    ws1.cell(srow, 4, n_art).font = Font(name=FONT, bold=True, size=10)
    ws1.cell(srow + 1, 1, "Total products:").font = Font(name=FONT, bold=True, size=10)
    ws1.cell(srow + 1, 4, len(prod_rows)).font = Font(name=FONT, bold=True, size=10)
    ws1.cell(srow + 3, 1,
             f"Source: Printful API v2 /sync-products, store {STORE}, {date.today().isoformat()}"
             ).font = Font(name=FONT, italic=True, size=9, color="808080")

    ws2 = wb.create_sheet("Variants")
    vh = ["Product", "Variant", "SKU", "Catalog Variant ID", "Retail Price", "Currency", "Technique", "Layer Types", "Artwork URL"]
    last2 = write_sheet(ws2, "Sync Variants Detail", vh, var_rows, [34, 44, 16, 16, 13, 9, 14, 18, 30])
    for r in range(4, last2 + 1):
        pc = ws2.cell(r, 5)
        if pc.value is not None:
            pc.number_format = '$#,##0.00;($#,##0.00);-'
            pc.alignment = Alignment(horizontal="right", vertical="center")
    return wb


def main():
    prod_rows, var_rows = fetch()
    wb = build(prod_rows, var_rows)
    wb.save(OUT)
    print(f"Saved {OUT}: {len(prod_rows)} products, {len(var_rows)} variants")


if __name__ == "__main__":
    main()
